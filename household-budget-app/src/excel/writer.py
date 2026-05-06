"""
Excel レポート生成モジュール

openpyxl を使い、月次支出レポートを Excel ファイルに書き出す。
シート構成:
  1. サマリー  - 収支サマリーとカテゴリ別支出円グラフ
  2. 明細      - 全取引の一覧
  3. カテゴリ別 - カテゴリ別月次推移グラフ
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from src.models.transaction import MonthlySummary, SavingsGoal, TransactionType


# カラーパレット
COLOR_HEADER = "1F4E79"
COLOR_INCOME = "70AD47"
COLOR_EXPENSE = "FF0000"
COLOR_ALT_ROW = "EBF3FB"
COLOR_BORDER = "BDD7EE"


class ExcelReportWriter:
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write(
        self,
        summary: MonthlySummary,
        analysis_text: str = "",
        goal: SavingsGoal | None = None,
    ) -> str:
        wb = Workbook()
        wb.remove(wb.active)  # デフォルトシートを削除

        self._write_summary_sheet(wb, summary, analysis_text, goal)
        self._write_detail_sheet(wb, summary)
        self._write_category_sheet(wb, summary)

        filename = self.output_dir / f"家計簿_{summary.year}{summary.month:02d}.xlsx"
        wb.save(filename)
        return str(filename)

    # ------------------------------------------------------------------ #
    #  サマリーシート                                                        #
    # ------------------------------------------------------------------ #

    def _write_summary_sheet(
        self,
        wb: Workbook,
        summary: MonthlySummary,
        analysis_text: str,
        goal: SavingsGoal | None = None,
    ) -> None:
        ws = wb.create_sheet("サマリー")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 50

        # タイトル
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = f"{summary.label} 家計簿レポート"
        title_cell.font = Font(name="游ゴシック", size=18, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # 生成日時
        ws["A2"] = f"生成日時: {datetime.now().strftime('%Y/%m/%d %H:%M')}"
        ws["A2"].font = Font(name="游ゴシック", size=9, color="888888")
        ws.row_dimensions[2].height = 16

        # 収支サマリー
        self._write_section_header(ws, "A4", "収支サマリー")
        rows = [
            ("収入合計", summary.total_income, COLOR_INCOME),
            ("支出合計", summary.total_expense, COLOR_EXPENSE),
            ("収支差額", summary.net, COLOR_INCOME if summary.net >= 0 else COLOR_EXPENSE),
        ]
        for i, (label, value, color) in enumerate(rows, start=5):
            row = i
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(name="游ゴシック", size=11, bold=True)
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = '#,##0"円"'
            ws[f"B{row}"].font = Font(name="游ゴシック", size=12, bold=True, color=color)
            ws[f"B{row}"].alignment = Alignment(horizontal="right")
            ws.row_dimensions[row].height = 22

        # 貯蓄目標ウィジェット
        goal_offset = 0
        if goal:
            goal_offset = self._write_goal_widget(ws, goal, summary, start_row=9)

        # カテゴリ別支出テーブル
        cat_start = 9 + goal_offset
        self._write_section_header(ws, f"A{cat_start}", "カテゴリ別支出")
        header_row = cat_start + 1
        ws[f"A{header_row}"] = "カテゴリ"
        ws[f"B{header_row}"] = "金額"
        ws[f"C{header_row}"] = "割合"
        for col in ["A", "B", "C"]:
            cell = ws[f"{col}{header_row}"]
            cell.font = Font(name="游ゴシック", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center")

        expense_cats = {k: v for k, v in summary.by_category.items() if v > 0 and k != "収入"}
        data_start = header_row + 1
        for i, (cat, amount) in enumerate(
            sorted(expense_cats.items(), key=lambda x: -x[1]), start=data_start
        ):
            fill = PatternFill("solid", fgColor=COLOR_ALT_ROW) if i % 2 == 0 else None
            ws[f"A{i}"] = cat
            ws[f"B{i}"] = amount
            ws[f"B{i}"].number_format = '#,##0"円"'
            ws[f"C{i}"] = amount / summary.total_expense if summary.total_expense else 0
            ws[f"C{i}"].number_format = "0.0%"
            for col in ["A", "B", "C"]:
                ws[f"{col}{i}"].font = Font(name="游ゴシック", size=10)
                if fill:
                    ws[f"{col}{i}"].fill = fill
            ws.row_dimensions[i].height = 18

        # 円グラフ (カテゴリ別支出)
        chart_row_start = data_start
        chart_row_end = chart_row_start + len(expense_cats) - 1
        if len(expense_cats) > 0:
            pie = PieChart()
            pie.title = "カテゴリ別支出"
            pie.style = 10
            pie.width = 14
            pie.height = 10
            labels = Reference(ws, min_col=1, min_row=chart_row_start, max_row=chart_row_end)
            data = Reference(ws, min_col=2, min_row=chart_row_start, max_row=chart_row_end)
            pie.add_data(data)
            pie.set_categories(labels)
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            ws.add_chart(pie, "D9")

        # AI フィードバック
        if analysis_text:
            feedback_row = max(chart_row_end + 3, data_start + len(expense_cats) + 2)
            self._write_section_header(ws, f"A{feedback_row}", "AI 支出分析・アドバイス")
            ws[f"A{feedback_row + 1}"] = analysis_text
            ws[f"A{feedback_row + 1}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws[f"A{feedback_row + 1}"].font = Font(name="游ゴシック", size=10)
            ws.merge_cells(
                f"A{feedback_row + 1}:D{feedback_row + max(analysis_text.count(chr(10)) + 2, 5)}"
            )
            ws.row_dimensions[feedback_row + 1].height = max(
                analysis_text.count("\n") * 15 + 30, 80
            )

    # ------------------------------------------------------------------ #
    #  明細シート                                                           #
    # ------------------------------------------------------------------ #

    def _write_detail_sheet(self, wb: Workbook, summary: MonthlySummary) -> None:
        ws = wb.create_sheet("取引明細")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        headers = ["日付", "説明", "カテゴリ", "金額", "収支", "データソース"]
        col_widths = [14, 35, 14, 16, 10, 20]
        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name="游ゴシック", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 22

        for i, tx in enumerate(summary.transactions, start=2):
            fill = PatternFill("solid", fgColor=COLOR_ALT_ROW) if i % 2 == 0 else None
            sign = "支出" if tx.is_expense else "収入"
            color = COLOR_EXPENSE if tx.is_expense else COLOR_INCOME
            row_data = [
                tx.date.strftime("%Y/%m/%d"),
                tx.description,
                tx.category.value,
                tx.amount,
                sign,
                tx.source,
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.font = Font(name="游ゴシック", size=10)
                if fill:
                    cell.fill = fill
                if col == 4:
                    cell.number_format = '#,##0"円"'
                    cell.alignment = Alignment(horizontal="right")
                if col == 5:
                    cell.font = Font(name="游ゴシック", size=10, color=color)
                    cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[i].height = 18

        # テーブルに変換してフィルター追加
        from openpyxl.worksheet.table import Table, TableStyleInfo
        last_row = len(summary.transactions) + 1
        if last_row > 1:
            tab = Table(
                displayName="TxTable",
                ref=f"A1:{get_column_letter(len(headers))}{last_row}",
            )
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(tab)

    # ------------------------------------------------------------------ #
    #  カテゴリ別シート                                                     #
    # ------------------------------------------------------------------ #

    def _write_category_sheet(self, wb: Workbook, summary: MonthlySummary) -> None:
        ws = wb.create_sheet("カテゴリ別集計")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18

        self._write_section_header(ws, "A1", "カテゴリ別支出集計")

        ws["A2"] = "カテゴリ"
        ws["B2"] = "合計金額"
        for cell in [ws["A2"], ws["B2"]]:
            cell.font = Font(name="游ゴシック", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center")

        expense_cats = {k: v for k, v in summary.by_category.items() if v > 0 and k != "収入"}
        sorted_cats = sorted(expense_cats.items(), key=lambda x: -x[1])
        for i, (cat, amount) in enumerate(sorted_cats, start=3):
            ws[f"A{i}"] = cat
            ws[f"B{i}"] = amount
            ws[f"B{i}"].number_format = '#,##0"円"'
            ws.row_dimensions[i].height = 18

        # 棒グラフ
        if sorted_cats:
            bar = BarChart()
            bar.type = "col"
            bar.title = f"{summary.label} カテゴリ別支出"
            bar.style = 10
            bar.y_axis.title = "金額 (円)"
            bar.width = 20
            bar.height = 12
            data = Reference(ws, min_col=2, min_row=2, max_row=2 + len(sorted_cats))
            cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(sorted_cats))
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            ws.add_chart(bar, "D2")

    # ------------------------------------------------------------------ #
    #  貯蓄目標ウィジェット                                                 #
    # ------------------------------------------------------------------ #

    def _write_goal_widget(
        self,
        ws,
        goal: SavingsGoal,
        summary: MonthlySummary,
        start_row: int,
    ) -> int:
        """貯蓄目標セクションを描画し、使用した行数を返す。"""
        self._write_section_header(ws, f"A{start_row}", "貯蓄目標 進捗")

        on_track = goal.on_track(summary.net)
        forecast = goal.forecast(summary.net)
        progress_pct = goal.progress_rate

        rows = [
            ("目標金額", f"{goal.target_amount:,}円", "2E75B6"),
            ("達成期限", goal.deadline.strftime("%Y年%m月末"), "2E75B6"),
            ("累計貯蓄額", f"{goal.current_savings:,}円", "2E75B6"),
            ("今月の貯蓄", f"{summary.net:,}円", COLOR_INCOME if summary.net >= 0 else COLOR_EXPENSE),
            ("必要月次貯蓄", f"{goal.required_monthly_savings:,}円", "2E75B6"),
            ("ペース判定", "✅ 目標ペース達成" if on_track else "⚠️ ペース不足", COLOR_INCOME if on_track else "FF8C00"),
            ("年末予測", f"{forecast:,}円 ({'達成見込み' if forecast >= goal.target_amount else '未達見込み'})",
             COLOR_INCOME if forecast >= goal.target_amount else COLOR_EXPENSE),
        ]
        if goal.stretch_amount:
            rows.append(
                ("ストレッチ目標", f"{goal.stretch_amount:,}円（月{goal.stretch_required_monthly:,}円必要）", "7030A0")
            )

        for i, (label, value, color) in enumerate(rows, start=start_row + 1):
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = Font(name="游ゴシック", size=10, bold=True)
            ws[f"A{i}"].fill = PatternFill("solid", fgColor="F2F2F2")
            ws[f"B{i}"] = value
            ws[f"B{i}"].font = Font(name="游ゴシック", size=10, bold=True, color=color)
            ws.row_dimensions[i].height = 20

        # プログレスバー（セルの背景色で表現）
        bar_row = start_row + len(rows) + 2
        ws[f"A{bar_row}"] = "達成率"
        ws[f"A{bar_row}"].font = Font(name="游ゴシック", size=10, bold=True)
        filled = max(int(progress_pct * 10), 0)
        for col_idx in range(1, 11):
            col = get_column_letter(col_idx + 1)
            cell = ws[f"{col}{bar_row}"]
            cell.fill = PatternFill(
                "solid", fgColor=COLOR_INCOME if col_idx <= filled else "D9D9D9"
            )
            ws.column_dimensions[col].width = 4
        pct_cell = ws[f"L{bar_row}"]
        pct_cell.value = f"{progress_pct:.1%}"
        pct_cell.font = Font(name="游ゴシック", size=10, bold=True, color=COLOR_INCOME)
        ws.row_dimensions[bar_row].height = 18

        return len(rows) + 4  # 使用した行数を返す

    # ------------------------------------------------------------------ #
    #  ユーティリティ                                                       #
    # ------------------------------------------------------------------ #

    def _write_section_header(self, ws, cell_ref: str, title: str) -> None:
        cell = ws[cell_ref]
        cell.value = title
        cell.font = Font(name="游ゴシック", size=12, bold=True, color=COLOR_HEADER)
        ws.row_dimensions[cell.row].height = 22
