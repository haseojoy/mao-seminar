import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import date

wb = openpyxl.Workbook()

# ── カラーパレット ──────────────────────────────────────
HEADER_BG   = "1F3864"   # 濃紺
HEADER_FG   = "FFFFFF"
CAT_BG      = "2E75B6"   # 中紺
CAT_FG      = "FFFFFF"
SUBCAT_BG   = "D6E4F0"   # 薄青
KNOWN_BG    = "FFFFFF"
UNKNOWN_BG  = "FFF2CC"   # 薄黄 = 要確認
ALERT_BG    = "FCE4D6"   # 薄オレンジ = 削れる
TOTAL_BG    = "E2EFDA"   # 薄緑 = 合計
BORDER_COL  = "AAAAAA"

thin = Side(style="thin", color=BORDER_COL)
thick= Side(style="medium", color="888888")

def border(t=False, b=False, l=False, r=False):
    return Border(
        top=thick if t else thin,
        bottom=thick if b else thin,
        left=thick if l else thin,
        right=thick if r else thin,
    )

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Meiryo UI")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ══════════════════════════════════════════════════════════════
# シート 1 : 月次支出サマリー
# ══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "月次支出サマリー"
ws.sheet_view.showGridLines = False

# 列幅
col_widths = {"A": 5, "B": 28, "C": 22, "D": 14, "E": 14, "F": 30}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# 行高
ws.row_dimensions[1].height = 40
ws.row_dimensions[2].height = 18

# タイトル
ws.merge_cells("A1:F1")
c = ws["A1"]
c.value = "支出管理シート  （2026年5月 作成）"
c.font = Font(bold=True, color=HEADER_FG, size=14, name="Meiryo UI")
c.fill = fill(HEADER_BG)
c.alignment = center()

# ヘッダー行
headers = ["#", "カテゴリ / サービス名", "支払方法 / 備考", "月額(円)", "年額(円)", "アクション・メモ"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = font(bold=True, color=CAT_FG)
    c.fill = fill(CAT_BG)
    c.alignment = center()
    c.border = border(t=True, b=True, l=(col==1), r=(col==6))

# ──────────────────────────────────────────────────────────────
# データ定義
# fmt: (no, name, method, monthly, annual, action, row_type)
# row_type: "cat"=カテゴリ見出し, "item"=通常, "check"=要確認, "cut"=削れる候補
# ──────────────────────────────────────────────────────────────
rows = [
    # ① 家賃
    ("cat", "① 家賃"),
    ("check", "家賃（ASSETIA 管理物件）", "エポスカード 家賃保証", None, None, "契約更新中・金額要確認 → カード明細参照"),
    ("space",),

    # ② 光熱費・通信費
    ("cat", "② 光熱費・通信費"),
    ("item",  "電気代（Japan電力）", "口座引落", None, None, "4月分請求済・金額はPDF参照"),
    ("check", "水道・ガス", "口座引落", None, None, "金額要確認"),
    ("check", "携帯・スマホ", "カード/口座", None, None, "金額要確認（ahamo/povo/LINEMO 等？）"),
    ("space",),

    # ③ 保険
    ("cat", "③ 保険"),
    ("check", "国民健康保険 / 社会保険", "口座引落", None, None, "金額要確認"),
    ("check", "生命保険・医療保険", "口座引落 / カード", None, None, "金額要確認"),
    ("check", "火災保険（賃貸）", "年払い等", None, None, "金額要確認"),
    ("space",),

    # ④ サブスク（確認済）
    ("cat", "④ サブスク（確認済）"),
    ("item",  "Google AI Pro 5TB（Google One）", "エポスカード", 2900, 34800, "2026/06/02 解約予定 → 解約済で削減"),
    ("item",  "Dropbox Simple 月間プラン", "Apple / iPhone", 900, 10800, "クラウドはGoogle Driveで代替可？要検討"),
    ("item",  "AWS（クラウド利用料）", "カード（業務用）", None, None, "請求書PDF参照・業務費として計上可"),
    ("cut",   "日経電子版（個人プラン）", "カード", None, None, "解約済（〜7/9まで無料継続）→ 削減完了"),
    ("space",),

    # ⑤ サブスク（要確認）
    ("cat", "⑤ サブスク（要確認）"),
    ("check", "Amazon Prime / Audible", "カード", None, None, "利用有無・金額を確認"),
    ("check", "U-NEXT", "カード", None, None, "過去利用履歴あり・現在契約状況確認"),
    ("check", "その他サブスク", "—", None, None, "カード明細を見て洗い出し"),
    ("space",),

    # ⑥ 車・バイク
    ("cat", "⑥ 車・バイク"),
    ("check", "駐車場 / バイク駐輪場", "口座引落 / カード", None, None, "金額要確認"),
    ("check", "自動車保険 / バイク保険", "年払い等", None, None, "金額要確認"),
    ("check", "ガソリン・車検・維持費", "現金 / カード", None, None, "月平均を算出する"),
    ("space",),

    # ⑦ 外食・コンビニ・Uber
    ("cat", "⑦ 外食・コンビニ・Uber（変動費）"),
    ("cut",   "外食・カフェ", "楽天カード / エポスカード", None, None, "カード明細で集計・削減優先"),
    ("cut",   "Uber Eats / 出前館", "カード", None, None, "カード明細で集計・削減優先"),
    ("cut",   "コンビニ", "現金 / カード", None, None, "月合計を集計・削減優先"),
    ("space",),

    # ⑧ 飲み会・交際費
    ("cat", "⑧ 飲み会・交際費（変動費）"),
    ("cut",   "飲み会・会食", "現金 / カード", None, None, "月合計を集計・削減優先"),
    ("cut",   "贈り物・冠婚葬祭", "現金 / カード", None, None, "突発的支出として記録"),
    ("space",),

    # ⑨ 美容・服・趣味
    ("cat", "⑨ 美容・服・趣味（変動費）"),
    ("item",  "コアラ皮膚科（5/2）", "M3デジカルスマート", 1520, None, "美容皮膚科・単発"),
    ("item",  "美容室（HotPepper Beauty経由）", "カード", None, None, "4/30 予約利用"),
    ("cut",   "服・ファッション", "カード", None, None, "カード明細で集計・削減優先"),
    ("cut",   "趣味・ゲーム・娯楽", "カード / 現金", None, None, "カード明細で集計・削減優先"),
    ("space",),

    # ⑩ 日用品・小さな浪費
    ("cat", "⑩ 日用品・小さな浪費（変動費）"),
    ("item",  "Baluko コインランドリー", "クレジットカード", None, None, "¥400(洗濯)+¥200(乾燥)/回 × 月複数回"),
    ("cut",   "Amazon 衝動買い", "カード", None, None, "カード明細で集計・削減優先"),
    ("cut",   "日用品雑貨", "現金 / カード", None, None, "カード明細で集計・削減優先"),
    ("cut",   "粗大ごみ手数料（5/30）", "世田谷区", None, None, "単発支出"),
    ("space",),

    # リボ・借金
    ("cat", "⑪ リボ・借金・分割払い"),
    ("check", "楽天カード リボ残高", "楽天カード(Visa)", None, None, "4月支払0円・残高・リボ状況を確認"),
    ("check", "エポスカード リボ / 分割", "エポスカード", None, None, "アプリ「お支払予定額照会」で確認"),
    ("check", "その他借入・ローン", "—", None, None, "カードローン残高も確認"),
]

row_idx = 3
item_no = 0
cat_rows = []   # (row, label) for summary chart

for entry in rows:
    if entry[0] == "space":
        ws.row_dimensions[row_idx].height = 6
        row_idx += 1
        continue

    if entry[0] == "cat":
        ws.row_dimensions[row_idx].height = 22
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        c = ws[f"A{row_idx}"]
        c.value = entry[1]
        c.font = font(bold=True, color=CAT_FG, size=10)
        c.fill = fill(CAT_BG)
        c.alignment = left()
        c.border = border(t=True, b=True, l=True, r=True)
        cat_rows.append((row_idx, entry[1]))
        row_idx += 1
        continue

    # item / check / cut
    kind  = entry[0]
    name  = entry[1]
    mthd  = entry[2]
    mon   = entry[3]
    ann   = entry[4] if len(entry) > 4 else (mon * 12 if mon else None)
    memo  = entry[5] if len(entry) > 5 else ""

    item_no += 1
    bg = UNKNOWN_BG if kind == "check" else (ALERT_BG if kind == "cut" else KNOWN_BG)
    ws.row_dimensions[row_idx].height = 18

    vals = [item_no, name, mthd, mon, ann, memo]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row_idx, column=col)
        c.value = v if v is not None else "要確認"
        c.fill = fill(bg)
        c.font = font(size=10)
        c.alignment = center() if col in (1, 4, 5) else left()
        c.border = border(l=(col==1), r=(col==6))
        if col in (4, 5) and v is not None:
            c.number_format = '#,##0'
    row_idx += 1

# 合計行
ws.row_dimensions[row_idx].height = 24
ws.merge_cells(f"A{row_idx}:C{row_idx}")
c = ws[f"A{row_idx}"]
c.value = "合計（確認済のみ）"
c.font = font(bold=True, size=11)
c.fill = fill(TOTAL_BG)
c.alignment = center()
c.border = border(t=True, b=True, l=True)

# 月額合計（D列）
total_row = row_idx
ws[f"D{row_idx}"].value = "=SUMIF(D3:D{},\"<>要確認\",D3:D{})".format(row_idx-1, row_idx-1)
ws[f"D{row_idx}"].number_format = '#,##0'
ws[f"D{row_idx}"].font = font(bold=True, size=11)
ws[f"D{row_idx}"].fill = fill(TOTAL_BG)
ws[f"D{row_idx}"].alignment = center()
ws[f"D{row_idx}"].border = border(t=True, b=True)

# 年額合計（E列）
ws[f"E{row_idx}"].value = "=SUMIF(E3:E{},\"<>要確認\",E3:E{})".format(row_idx-1, row_idx-1)
ws[f"E{row_idx}"].number_format = '#,##0'
ws[f"E{row_idx}"].font = font(bold=True, size=11)
ws[f"E{row_idx}"].fill = fill(TOTAL_BG)
ws[f"E{row_idx}"].alignment = center()
ws[f"E{row_idx}"].border = border(t=True, b=True)

ws[f"F{row_idx}"].fill = fill(TOTAL_BG)
ws[f"F{row_idx}"].border = border(t=True, b=True, r=True)

# 凡例
row_idx += 2
legend = [
    ("  白背景 = 金額確認済", KNOWN_BG),
    ("  黄背景 = 要確認（カード明細参照）", UNKNOWN_BG),
    ("  オレンジ = 削減優先候補", ALERT_BG),
]
for label, bg in legend:
    ws.row_dimensions[row_idx].height = 16
    ws.merge_cells(f"A{row_idx}:F{row_idx}")
    c = ws[f"A{row_idx}"]
    c.value = label
    c.fill = fill(bg)
    c.font = font(size=9, color="444444")
    c.alignment = left()
    row_idx += 1

# ══════════════════════════════════════════════════════════════
# シート 2 : カード明細チェックリスト
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("カード明細チェックリスト")
ws2.sheet_view.showGridLines = False

col_widths2 = {"A": 6, "B": 22, "C": 32, "D": 14, "E": 16, "F": 14, "G": 24}
for col, w in col_widths2.items():
    ws2.column_dimensions[col].width = w

ws2.row_dimensions[1].height = 36
ws2.merge_cells("A1:G1")
c = ws2["A1"]
c.value = "カード明細チェックリスト  ― 月次振り返り用"
c.font = Font(bold=True, color=HEADER_FG, size=13, name="Meiryo UI")
c.fill = fill(HEADER_BG)
c.alignment = center()

headers2 = ["#", "カード名", "利用店舗 / サービス名", "金額(円)", "カテゴリ", "削減対象?", "メモ"]
ws2.row_dimensions[2].height = 20
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = font(bold=True, color=CAT_FG)
    c.fill = fill(CAT_BG)
    c.alignment = center()

# 入力例行
examples = [
    ("楽天カード(Visa)", "（例）Uber Eats", 3500, "外食・Uber", "Yes", "月3回合計"),
    ("楽天カード(Visa)", "（例）コンビニ各社", 4200, "日用品・浪費", "Yes", "週平均1000円"),
    ("エポスカード",     "（例）Amazon.co.jp", 6800, "日用品・趣味", "要判断", "衝動買い含む"),
    ("エポスカード",     "Google AI Pro 5TB", 2900, "サブスク", "解約済", "6/2解約"),
    ("エポスカード",     "Dropbox Simple",    900,  "サブスク", "検討中", "代替案あり"),
    ("Apple/iPhone",    "Dropbox Simple",    900,  "サブスク", "検討中", "Appleからも請求"),
]
for i, (card, shop, amt, cat, cut, memo) in enumerate(examples, 1):
    r = i + 2
    ws2.row_dimensions[r].height = 18
    bg = ALERT_BG if cut in ("Yes", "解約済") else (UNKNOWN_BG if cut == "検討中" else KNOWN_BG)
    for col, v in enumerate([i, card, shop, amt, cat, cut, memo], 1):
        c = ws2.cell(row=r, column=col, value=v)
        c.fill = fill(bg)
        c.font = font(size=10, color="666666")
        c.alignment = center() if col in (1, 4, 6) else left()
        if col == 4:
            c.number_format = '#,##0'

# 空白入力行（20行）
for i in range(7, 27):
    ws2.row_dimensions[i].height = 18
    for col in range(1, 8):
        c = ws2.cell(row=i, column=col)
        c.border = border()
        c.fill = fill("F8F8F8")

# ══════════════════════════════════════════════════════════════
# シート 3 : 削減アクションプラン
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("削減アクションプラン")
ws3.sheet_view.showGridLines = False
col_widths3 = {"A": 5, "B": 26, "C": 14, "D": 14, "E": 14, "F": 14, "G": 28}
for col, w in col_widths3.items():
    ws3.column_dimensions[col].width = w

ws3.row_dimensions[1].height = 36
ws3.merge_cells("A1:G1")
c = ws3["A1"]
c.value = "削減アクションプラン"
c.font = Font(bold=True, color=HEADER_FG, size=13, name="Meiryo UI")
c.fill = fill(HEADER_BG)
c.alignment = center()

headers3 = ["#", "削減項目", "現在の月額(円)", "目標月額(円)", "月削減額(円)", "ステータス", "アクション内容"]
ws3.row_dimensions[2].height = 20
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=col, value=h)
    c.font = font(bold=True, color=CAT_FG)
    c.fill = fill(CAT_BG)
    c.alignment = center()

actions = [
    ("Google AI Pro (解約済)", 2900, 0, 2900, "完了", "2026/06/02解約済 → Google One Free へ"),
    ("日経電子版（解約済）", None, 0, None, "完了", "解約受付済（7/9まで継続）"),
    ("Dropbox（代替検討）", 900, 0, 900, "検討中", "Google Drive無料枠 or 解約を検討"),
    ("外食・Uber・コンビニ削減", None, None, None, "要集計", "カード明細を集計 → 月3万円を目標"),
    ("リボ・分割払い返済計画", None, None, None, "要確認", "楽天・エポスのリボ残高確認→繰上返済"),
    ("サブスク棚卸し（全体）", None, None, None, "今週中", "カード明細全件をカテゴリ分けして整理"),
    ("飲み会・交際費の見直し", None, None, None, "要集計", "月上限を決めて超えた分は翌月繰越し"),
    ("美容・服の予算化", None, None, None, "要集計", "月予算を設定してオーバーしない管理"),
]
status_colors = {"完了": "E2EFDA", "検討中": "FFF2CC", "要確認": ALERT_BG, "今週中": ALERT_BG, "要集計": "FFE6CC"}

for i, (item, cur, tgt, save, status, action) in enumerate(actions, 1):
    r = i + 2
    ws3.row_dimensions[r].height = 20
    bg = status_colors.get(status, KNOWN_BG)
    vals = [i, item, cur, tgt, save, status, action]
    for col, v in enumerate(vals, 1):
        c = ws3.cell(row=r, column=col)
        c.value = "—" if v is None else v
        c.fill = fill(bg)
        c.font = font(size=10)
        c.alignment = center() if col in (1, 3, 4, 5, 6) else left()
        c.border = border()
        if col in (3, 4, 5) and isinstance(v, (int, float)):
            c.number_format = '#,##0'

# 合計行
r_total = len(actions) + 3
ws3.row_dimensions[r_total].height = 22
ws3.merge_cells(f"A{r_total}:B{r_total}")
c = ws3[f"A{r_total}"]
c.value = "削減合計（確認済）"
c.font = font(bold=True, size=11)
c.fill = fill(TOTAL_BG)
c.alignment = center()
c.border = border(t=True, b=True, l=True)
for col in range(3, 8):
    c = ws3.cell(row=r_total, column=col)
    c.fill = fill(TOTAL_BG)
    c.border = border(t=True, b=True, r=(col==7))
    if col == 5:
        c.value = f"=SUMIF(E3:E{r_total-1},\"<>—\",E3:E{r_total-1})"
        c.number_format = '#,##0'
        c.font = font(bold=True, size=11)
        c.alignment = center()

# ══════════════════════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════════════════════
out = "/home/user/mao-seminar/支出管理シート_2026年5月.xlsx"
wb.save(out)
print(f"Saved: {out}")
